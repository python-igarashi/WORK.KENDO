# coding: cp932
import Defines
import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Side, Font
from openpyxl.utils import get_column_letter


# ─────────────────────────────
#  大会プログラム冊子の選手一覧作成用クラス
# ─────────────────────────────
class booklet:
	def __init__(self, player_num, summary_name):
		self.player_num = player_num
		self.summary_name = summary_name
		self.team_index = -1
		self.current_row = []
		self.summary = []
		self.teams_per_line = 3 # 一覧で横に並べるチーム数
	
	# ─────────────────────────────
	#  チームデータを追加する
	# ─────────────────────────────
	def append_team(self, team):
		self.team_index += 1
		
		team_offset_Y = self.team_index // self.teams_per_line
		team_offset_X = self.team_index % self.teams_per_line
		
		# 1チームあたりの行数列数を取得
		num_rows = 1 + 1 + (self.player_num * 2) # 行数 = 1[団体] + 1[監督] + ({選手数} * 2[選手ふりがな + 選手氏名])
		num_cols = 3 # ポジション + 名前 + 学年|段位
		
		# 1チームの内容保存用行列を初期化
		contents = [[""] * num_cols for i in range(num_rows)] # Empty文字の2次元配列を作成する
		
		# 一覧の左端のチームの場合、summaryに横に並べるチーム数分の行列を追加する
		if team_offset_X == 0:
			self.summary += [[""] * num_cols * self.teams_per_line for i in range(num_rows)]
		
		summary_offset_Y = team_offset_Y * num_rows
		summary_offset_X = team_offset_X * num_cols
		
		# 団体行
		row_cur = 0
		self.summary[row_cur + summary_offset_Y][0 + summary_offset_X] = "団体名"
		self.summary[row_cur + summary_offset_Y][1 + summary_offset_X] = Defines.get_booklet_groupname(team[0][0], self.summary_name)
		
		# 監督行
		row_cur += 1
		self.summary[row_cur + summary_offset_Y][0 + summary_offset_X] = "監督"
		self.summary[row_cur + summary_offset_Y][1 + summary_offset_X] = team[1][2]
		
		# 選手行
		for i in range(self.player_num):
			# ふりがな行
			row_cur += 1
			self.summary[row_cur + summary_offset_Y][0 + summary_offset_X] = team[2 + i][1] # ポジション(先鋒, ...)
			self.summary[row_cur + summary_offset_Y][1 + summary_offset_X] = team[2 + i][3] # ふりがな
			if self.summary_name.find("一般") >= 0:
				self.summary[row_cur + summary_offset_Y][2 + summary_offset_X] = team[2 + i][7] # 段位
			else:
				self.summary[row_cur + summary_offset_Y][2 + summary_offset_X] = team[2 + i][5] + ("" if team[2 + i][5] == "" else "年") # 学年 ("" if self.summary_name.find("一般") >= 0 else "年")  # 学年/段位
			
			# 氏名行
			row_cur += 1
			self.summary[row_cur + summary_offset_Y][1 + summary_offset_X] = team[2 + i][2] # 氏名(先鋒, ...)
	
	
	# ─────────────────────────────
	#  全てのチームをTSV形式で出力する
	# ─────────────────────────────
	def output_tsv(self, path):
		import csv
		with open(path, mode="w", newline="", encoding="utf-8") as f:
			writer = csv.writer(f, delimiter="\t")
			writer.writerows(self.summary)
	
	
	# ─────────────────────────────
	#  全てのチームをEXCEL形式で出力する
	# ─────────────────────────────
	def output_xlsx(self, path, init_workbook = False):
		# ─────────────────────────────
		#  Excel 初期化
		# ─────────────────────────────
		if init_workbook:
			# 既存のファイルを削除し、新しいエクセルファイルを作成する
			try:
				os.remove(path)
				print(f"[{path}] を削除しました。")
			except FileNotFoundError:
				pass
			wb = Workbook()

			# 新しいシート{self.summary_name}を作成し、それ以外のデフォルトのシートを削除する
			ws = wb.create_sheet(self.summary_name)
			for sheet_name in wb.sheetnames:
				if sheet_name != self.summary_name:
					wb.remove(wb[sheet_name])
			wb.save(path)
			print(f"[{path}] を作成しました。")

		else:
			# 既存のエクセルファイルを開く
			wb = load_workbook(path)

			# 既存のシート{self.summary_name}を削除し、新しいシート{self.summary_name}を作成する
			if self.summary_name in  wb.sheetnames:
				wb.remove(wb[self.summary_name])
			ws = wb.create_sheet(self.summary_name)
			wb.save(path)

		mergin = 1 # 1行目、1列目はそれぞれマージンエリアとし、EXCELのセル番号指定時にこの値を調整する

		ws.title = self.summary_name
		for team_idx in range(self.teams_per_line):
			ws.column_dimensions[get_column_letter(3 * team_idx + 1 + mergin)].width = ws.column_dimensions[get_column_letter(3 * team_idx + 1 + mergin)].width * 0.7 # ポジション の列幅
			ws.column_dimensions[get_column_letter(3 * team_idx + 2 + mergin)].width = ws.column_dimensions[get_column_letter(3 * team_idx + 2 + mergin)].width * 2   # 選手名称 の列幅
			ws.column_dimensions[get_column_letter(3 * team_idx + 3 + mergin)].width = ws.column_dimensions[get_column_letter(3 * team_idx + 3 + mergin)].width * 0.5 # 学年 の列幅

		thin = Side(style="thin")    # 細線
		thick = Side(style="double") # 太線
		
		
		# ─────────────────────────────
		#  チーム情報
		# ─────────────────────────────
		for row, rowdata in enumerate(self.summary, start=1):
			for col, value in enumerate(rowdata, start=1):
				ws.cell(row=row + mergin, column=col + mergin, value=value)
		
		
		# ─────────────────────────────
		#  ポジション/学年/段位 のセル結合
		# ─────────────────────────────
		for row, rowdata in enumerate(self.summary, start=1):
			for col, value in enumerate(rowdata, start=1):
				# 団体名を太字にする
				#if col % 3 == 1 and value != "" and value == "団体名":
				#	ws.cell(row + mergin, col + 1 + mergin).font = Font(bold=True)
				
				# ポジション
				if col % 3 == 1 and value != "" and value != "団体名" and value != "監督":
					ws.merge_cells(start_row=row + mergin, start_column=col + mergin, end_row=row+1 + mergin, end_column=col + mergin)
					ws.cell(row + mergin, col + mergin).alignment = Alignment(horizontal="right", vertical="center")
				
				# 学年/段位
				if col % 3 == 0 and value != "":
					ws.merge_cells(start_row=row + mergin, start_column=col + mergin, end_row=row+1 + mergin, end_column=col + mergin)
					ws.cell(row + mergin, col + mergin).alignment = Alignment(horizontal="left", vertical="center")
		
		
		# ─────────────────────────────
		#  チームの周囲に太線、それ以外に細線
		# ─────────────────────────────
		for row, rowdata in enumerate(self.summary, start=1):
			for team_idx in range(self.teams_per_line):
				if rowdata[team_idx * 3] == "団体名":
					row_bottom = row + (self.player_num * 2) + 1 # チームの最終行
					col = (team_idx * 3) + 1 # チームの最左列
					col_right = col + 2      # チームの最右列
					
					# 団体名の行（チームの一番上）
					ws.cell(row + mergin, col + mergin).border = Border(top=thick, bottom=thin, left=thick, right=thin)
					ws.cell(row + mergin, col + 1 + mergin).border = Border(top=thick, bottom=thin, left=thin)
					ws.cell(row + mergin, col + 2 + mergin).border = Border(top=thick, bottom=thin, right=thick)
					
					# 監督の行
					ws.cell(row + 1 + mergin, col + mergin).alignment = Alignment(horizontal="right")
					ws.cell(row + 1 + mergin, col + mergin).border = Border(top=thin, bottom=thin, left=thick, right=thin)
					ws.cell(row + 1 + mergin, col + 1 + mergin).border = Border(top=thin, bottom=thin, left=thin)
					ws.cell(row + 1 + mergin, col + 2 + mergin).border = Border(top=thin, bottom=thin, right=thick)
					
					# 選手毎
					for row2 in range(row + 2, row_bottom + 1, 2):
						ws.cell(row2 + mergin, col + mergin).border = Border(top=thin, left=thick, right=thin)
						ws.cell(row2 + mergin, col + 1 + mergin).border = Border(top=thin, left=thin, right=thin)
						ws.cell(row2 + mergin, col + 2 + mergin).border = Border(top=thin, left=thin, right=thick)
						
						ws.cell(row2 + 1 + mergin, col + mergin).border = Border(bottom=thin, left=thick, right=thin)
						ws.cell(row2 + 1 + mergin, col + 1 + mergin).border = Border(bottom=thin, left=thin, right=thin)
						ws.cell(row2 + 1 + mergin, col + 2 + mergin).border = Border(bottom=thin, left=thin, right=thick)
					
					# チームの一番下
					ws.cell(row_bottom + mergin, col + mergin).border = Border(bottom=thick, left=thick, right=thin)
					ws.cell(row_bottom + mergin, col + 1 + mergin).border = Border(bottom=thick, left=thin, right=thin)
					ws.cell(row_bottom + mergin, col + 2 + mergin).border = Border(bottom=thick, left=thin, right=thick)
		
		
		# ファイルを保存する
		wb.save(path)
