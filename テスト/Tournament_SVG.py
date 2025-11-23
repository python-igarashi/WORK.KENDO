from __future__ import annotations
from dataclasses import dataclass
from collections import defaultdict
from typing import List, Tuple, Optional, Union
import math, random, os

# --------------------
# 基本データ構造
# --------------------
@dataclass(frozen=True)
class Participant:
    name: str
    kana: str # ふりがな
    team: Optional[str]  # 団体名（なしなら None）

@dataclass(frozen=True)
class Ref:
    """前ラウンドの '試合番号' を参照（その勝者がここに入る）"""
    match_index: int  # 0-based within that round

Slot = Union[Participant, Ref]

@dataclass
class Match:
    left: Slot
    right: Slot
    winner: Optional[Participant] = None  # BYE確定などで分かる場合のみ埋める

Round = List[Match]

# --------------------
# ユーティリティ
# --------------------
def next_power_of_two(n: int) -> int:
    return 1 if n <= 1 else 2 ** math.ceil(math.log2(n))

def bye_positions_balanced_halves(M: int, B: int) -> list[int]:
    """
    総試合数 M を 4 クォーターに分けつつ、
    まず Top(=Q1+Q2) / Bottom(=Q3+Q4) で BYE を均等配分し、
    その後 各半分の2クォーターへ均等配分して配置位置(0..M-1)を返す。
    """
    if B <= 0:
        return []

    # 1) クォーターサイズ（Mを4等分、端数は上から配分）
    base = M // 4
    rem  = M % 4
    q_sizes = [(base + (1 if i < rem else 0)) for i in range(4)]  # [Q1,Q2,Q3,Q4]

    # 2) 半分(TOP/BOTTOM)でBYE数を均等化
    #   half_byes[0]=Top, half_byes[1]=Bottom
    half_byes = [B // 2, B // 2]
    if B % 2 == 1:
        # 端数は、クォーター合計の大きい方(Top or Bottom)に付与
        top_size = q_sizes[0] + q_sizes[1]
        bot_size = q_sizes[2] + q_sizes[3]
        if top_size >= bot_size:
            half_byes[0] += 1
        else:
            half_byes[1] += 1

    # 3) 各半分の2クォーターへ均等配分（端数は“広い方”のクォーターへ）
    q_byes = [0, 0, 0, 0]  # [Q1,Q2,Q3,Q4]

    def split_into_two(total_byes: int, size_a: int, size_b: int) -> tuple[int, int]:
        a = total_byes // 2
        b = total_byes - a
        # 端数配分の調整：広い方に多めを寄せる
        if size_a < size_b and a > b:
            a, b = b, a
        if size_b < size_a and b > a:
            a, b = b, a
        return a, b

    q_byes[0], q_byes[1] = split_into_two(half_byes[0], q_sizes[0], q_sizes[1])  # Top -> Q1,Q2
    q_byes[2], q_byes[3] = split_into_two(half_byes[1], q_sizes[2], q_sizes[3])  # Bottom -> Q3,Q4

    # 4) 各クォーター内で等間隔の試合に配置 → グローバルindexへ変換
    res = []
    offset = 0
    for q in range(4):
        size = q_sizes[q]
        b    = q_byes[q]
        if size == 0 or b == 0:
            # 何も置かない
            pass
        else:
            # 等間隔配置（重複時は近傍にずらす）
            local = []
            used = set()
            for i in range(b):
                idx = round((i + 0.5) * size / b) - 1
                idx = max(0, min(size - 1, idx))
                if idx in used:
                    # 近い空きを探す
                    found = None
                    for d in range(1, size):
                        for cand in (idx - d, idx + d):
                            if 0 <= cand < size and cand not in used:
                                found = cand
                                break
                        if found is not None:
                            break
                    idx = found if found is not None else idx
                used.add(idx)
                local.append(idx)
            res.extend([offset + x for x in sorted(local)])
        offset += size

    return sorted(res)

# --------------------
# “同団体回避”を意識した並べ方（バケット→ラウンドロビン）
# --------------------
def diversified_order(participants: List[Participant], seed: Optional[int]) -> List[Participant]:
    """チーム分散しつつ、並びの規則性を弱める"""
    import random
    rnd = random.Random(seed)

    # バケット化 & 各バケット内はシャッフル
    buckets = defaultdict(list)
    for p in participants:
        buckets[p.team].append(p)
    for k in buckets:
        rnd.shuffle(buckets[k])

    # ① チームを「人数降順」でグルーピングし、同サイズ内はランダム
    by_size = defaultdict(list)
    for team, members in buckets.items():
        by_size[len(members)].append(team)

    sizes = sorted(by_size.keys(), reverse=True)
    bucket_keys: List[Optional[str]] = []
    for s in sizes:
        ks = by_size[s][:]
        rnd.shuffle(ks)  # ← 同規模チームの順序をランダム化
        bucket_keys.extend(ks)

    # ② 先頭の開始オフセットもランダムに回転
    if bucket_keys:
        off = rnd.randrange(len(bucket_keys))
        bucket_keys = bucket_keys[off:] + bucket_keys[:off]

    # ③ 取り出し周回ごとに左右反転（serpentine）して規則性を崩す
    order: List[Participant] = []
    forward = True
    while any(buckets[k] for k in bucket_keys):
        keys = bucket_keys if forward else list(reversed(bucket_keys))
        for k in keys:
            if buckets[k]:
                order.append(buckets[k].pop())
        forward = not forward

    # --- 上下偏り軽減のため、最終的な並びを上下で交互ミックス ---
    mid = len(order) // 2
    top = order[:mid]
    bottom = order[mid:]
    mixed = []
    for i in range(max(len(top), len(bottom))):
        if i < len(top):
            mixed.append(top[i])
        if i < len(bottom):
            mixed.append(bottom[i])

    return mixed

# --------------------
# 1回戦：四分割均等BYE割当（BYE vs BYE なし）
# --------------------
Pair = Tuple[Participant, Participant]

def _assign_pairs_with_bye_and_lookahead(
    order: list[Participant],
    M: int,
    bye_pos: set[int],
    lookahead: int,
    rnd=None,
) -> list[tuple[Participant, Participant]]:
    """
    1回戦の割付を“親ペア（2試合）単位”で行い、
    親が BYE×2 の場合は 2回戦で同一団体同士になりにくいように
    異なる団体の選手を優先して BYE に割り当てる。
    それ以外は従来の先読みロジックを維持。
    """
    if rnd is None:
        import random
        rnd = random.Random()

    bye = Participant("BYE", None, None)
    order = order[:]  # 破壊しない
    result: list[tuple[Participant, Participant] | None] = [None] * M

    idx = 0

    def pop_next() -> Participant:
        nonlocal idx
        p = order[idx] if idx < len(order) else bye
        if idx < len(order):
            idx += 1
        return p

    def pick_with_avoid(team_to_avoid: Optional[str]) -> Participant:
        """先読み窓から team_to_avoid を避けて1人選ぶ（なければ順番どおり）"""
        nonlocal idx
        w_end = min(idx + lookahead, len(order))
        if team_to_avoid:
            cands = [j for j in range(idx, w_end)
                     if not (order[j].team and order[j].team == team_to_avoid)]
            if cands:
                j = rnd.choice(cands)
                p = order[j]
                for k in range(j, idx, -1):
                    order[k] = order[k - 1]
                order[idx] = p
                idx += 1
                return p
        return pop_next()

    def _pair_fill_order(M: int, rnd) -> list[int]:
        """親ペア（2試合）を上↔下の順で埋めるための開始インデックス列を返す。
       例: M=8(=4ペア) → [0, 6, 2, 4]（= (0,1), (6,7), (2,3), (4,5)）"""
        starts = list(range(0, M, 2))
        seq = []
        i, j = 0, len(starts) - 1
        while i <= j:
            seq.append(starts[i])
            if i != j:
                seq.append(starts[j])
            i += 1
            j -= 1
        # ほんの少しだけランダム化（固定化回避）
        if len(seq) > 1:
            r = rnd.randrange(len(seq))
            seq = seq[r:] + seq[:r]
            if rnd.random() < 0.5:
                seq.reverse()
        return seq

    # 親ペア（2試合単位）で処理
    for k in _pair_fill_order(M, rnd):
        i0, i1 = k, k + 1

        both_bye = (i0 in bye_pos) and (i1 in bye_pos)
        if both_bye:
            # 2回戦同団体を避けるよう、異なる団体の2人を優先して選ぶ
            a = pop_next()
            b = pick_with_avoid(a.team)

            # 両方BYEの2試合を埋める
            result[i0] = (a, bye)
            result[i1] = (b, bye)
            continue

        # 片側だけBYE or BYEなし → 従来ロジック（各試合独立）
        for i in (i0, i1):
            if i >= M:  # 奇数保険
                break
            if result[i] is not None:
                continue

            if i in bye_pos:
                a = pop_next()
                result[i] = (a, bye)
                continue

            # 非BYE試合：a を出して、b は先読みで a と同団体を避けて選ぶ
            a = pop_next()
            b = pick_with_avoid(a.team)
            result[i] = (a, b)

    # None の保険処理
    for i in range(M):
        if result[i] is None:
            result[i] = (bye, bye)
    return result

def _conflicts(pairs: List[Pair]) -> int:
    """
    衝突スコアを返す。
    - 1回戦の同一団体対戦: +1
    - 2回戦の同一団体対戦: +1   ※(BYE,P) vs (BYE,Q) で, P と Q が同一団体: +3（2回戦で直接同一団体）
    """
    def is_bye(x) -> bool:
        return (x is None) or (getattr(x, "name", None) == "BYE")

    v1 = 0 # 第一優先チェック
    v2 = 0 # 第二優先チェック

    # 1) 1回戦の同一団体 → 可能な限り避ける
    for a, b in pairs:
        if not is_bye(a) and not is_bye(b):
            ta = getattr(a, "team", None)
            tb = getattr(b, "team", None)
            if ta and tb and ta == tb:
                v1 += 1

    # 2) 2回戦が初戦の衝突（親ノード=2試合ごと）
    for i in range(0, len(pairs), 2):
        (aL, aR) = pairs[i]
        (bL, bR) = pairs[i+1]

        a_real = [x for x in (aL, aR) if not is_bye(x)]
        b_real = [x for x in (bL, bR) if not is_bye(x)]

        a_has_bye = (len(a_real) == 1)
        b_has_bye = (len(b_real) == 1)

        # (BYE,P) vs (BYE,Q): P と Q は2回戦で直接対戦 → 可能な限り避ける
        if a_has_bye and b_has_bye and a_real and b_real:
            tP, tQ = getattr(a_real[0], "team", None), getattr(b_real[0], "team", None)
            if tP and tQ and tP == tQ:
                v1 += 1

        # (BYE,P) vs (Q1,Q2): P は Q1/Q2 のどちらとも同一団体NG → 軽く避ける（10）
        # (P1,P2) vs (Q1,Q2): P1/P2 は Q1/Q2 のどちらとも同一団体NG → 軽く避けるが、ある程度やむを得ない事とする（1）
        else:
            for P in a_real:
                for Q in b_real:
                    tP, tQ = getattr(P, "team", None), getattr(Q, "team", None)
                    if tP and tQ and tP == tQ:
                        if a_has_bye or b_has_bye:
                            v2 += 10
                        else:
                            v2 += 1

    return (v1, v2)

def make_first_round_pairs_quarter_even(
    participants: List[Participant],
    seed: Optional[int] = None,
    restarts: int = 1000,     # 複数リスタート
    lookahead: int = 6,      # 先読み窓幅（団体衝突を回避）
    max_local_iters: int = 2000,
) -> List[Pair]:
    """
    四分割BYE配分を維持しながら、同団体の初戦を強く回避。
    - diversified_order のランダム性は活かす
    - 先読み割付 + 強化ローカル最適化
    - 最良案（同団体カード数最小）を採用
    """
    N = len(participants)
    T = next_power_of_two(N)
    M = T // 2
    B = T - N
    bye_pos = set(bye_positions_balanced_halves(M, B))

    import random
    base_rnd = random.Random(seed)
    best_pairs = None
    best_c1 = 10**9
    best_c2 = 10**9

    for r in range(max(1, restarts)):
        # 乱数をずらす（seed未指定なら毎回異なる）
        sub_seed = base_rnd.getrandbits(64)
        order = diversified_order(participants, sub_seed)

        # 1) 先読みで割付
        rnd = random.Random(sub_seed)
        pairs = _assign_pairs_with_bye_and_lookahead(order, M, bye_pos, lookahead, rnd)

        c1, c2 = _conflicts(pairs)
        if c1 < best_c1:
            best_pairs, best_c1, best_c2 = pairs, c1, c2
            if best_c1 == 0 and best_c2 == 0:   # これ以上はない
                break
        elif c1 == best_c1 and c2 < best_c2:
            best_pairs, best_c1, best_c2 = pairs, c1, c2
            if best_c1 == 0 and best_c2 == 0:   # これ以上はない
                break

    return best_pairs if best_pairs is not None else []

# --------------------
# ブラケット構築（決勝まで）
# --------------------
def build_full_bracket(participants: List[Participant], seed: Optional[int] = None) -> List[Round]:
    first_pairs = make_first_round_pairs_quarter_even(participants, seed=seed, restarts=1000)
    rounds: List[Round] = []

    r0: Round = []
    for (a, b) in first_pairs:
        m = Match(left=a, right=b, winner=None)
        # BYE自動処理（1回戦のみ）
        if a.name == "BYE" and b.name != "BYE":
            m.winner = b
        elif b.name == "BYE" and a.name != "BYE":
            m.winner = a
        elif a.name == "BYE" and b.name == "BYE":
            # ここは起きない想定（禁止済み）だが保険で右勝者
            m.winner = b
        r0.append(m)
    rounds.append(r0)

    current_round = r0
    while len(current_round) > 1:
        next_round: Round = []
        for i in range(0, len(current_round), 2):
            L, R = current_round[i], current_round[i + 1]
            left_slot  = L.winner if L.winner else Ref(i)
            right_slot = R.winner if R.winner else Ref(i + 1)
            next_round.append(Match(left=left_slot, right=right_slot, winner=None))
        rounds.append(next_round)
        current_round = next_round
    return rounds

# ----------------------------
# コンソール出力ユーティリティ
# ----------------------------
def print_bracket(rounds: List[Round]) -> None:
    def round_name(idx: int, total_rounds: int) -> str:
        rnum = idx + 1
        if rnum == total_rounds: return "Final"
        if rnum == total_rounds - 1: return "Semifinal"
        if rnum == total_rounds - 2: return "Quarterfinal"
        return f"Round {rnum}"

    def slot_str(slot: Slot) -> str:
        if isinstance(slot, Participant):
            return f"{slot.name} ({slot.team or '-'})"
        return f"Winner of Match {slot.match_index + 1}"

    total = len(rounds)
    for r_idx, rnd in enumerate(rounds):
        #print(f"\n=== {round_name(r_idx, total)} ===")
        for m_idx, m in enumerate(rnd, start=1):
            left = slot_str(m.left)
            right = slot_str(m.right)
            w = f" -> WINNER: {m.winner.name}" if m.winner else ""
            print(f"Match {m_idx}: {left}  vs  {right}{w}")
        break

# ------------------------------
# グラフィカル出力ユーティリティ
# ------------------------------
# 既存の Participant, Ref, Match, Round, build_full_bracket, slot_str がある前提です。
from typing import List, Tuple, Union, Optional
from dataclasses import dataclass
import xml.sax.saxutils as xml_escape


# --- SVG 出力 ---
def save_bracket_svg(
    rounds: List[Round],
    path: str = "bracket.svg",
    cell_w: int = 300,
    cell_h: int = 28,
    unit_v: int = 34,
    h_gap: int = 40,         # 1回戦→2回戦の基準ギャップ
    compact_gap: int = 60,   # 2回戦以降の基準ギャップ
    min_seg: int = 26,       # 各水平セグメントの最小長(px)
    font_size: int = 12,
    margin: int = 20,
    font_family: str = "monospace", # "monospace", "MS Mincho", "MS Gothic"
    print_match_no: bool = True,
    match_name = "xxxxxxxxの部",
    match_date = "2099.12.31",
    match_place1 = "第一試合場",
    match_place2 = "第二試合場",
) -> None:
    import xml.sax.saxutils as xml_escape

    # --- helpers ------------------------------------------------------------
    def is_real(x: Slot) -> bool:
        return isinstance(x, Participant) and x.name != "BYE"

    def assign_match_numbers(rounds: List[Round]) -> dict[tuple[int, int], int]:
        """BYEを除く全ラウンドの実施試合に通し番号を振る。"""
        n = 1
        mapping = {}
        for r, rnd in enumerate(rounds):
            for i, m in enumerate(rnd):
                # 左右のどちらも BYE でない（＝実際に行われる試合）
                left_is_bye  = isinstance(m.left, Participant)  and m.left.name  == "BYE"
                right_is_bye = isinstance(m.right, Participant) and m.right.name == "BYE"
                if not (left_is_bye or right_is_bye):
                    mapping[(r, i)] = n
                    n += 1
        return mapping

    esc = xml_escape.escape
    R = len(rounds)
    if R == 0:
        with open(path, "w", encoding="utf-8") as f:
            f.write('<svg xmlns="http://www.w3.org/2000/svg" width="200" height="80"/>')
        return

    # --- 番号付け用マップ作成（1回戦の選手） ------------------------
    player_numbering: dict[Participant, int] = {}
    counter = 1
    r0 = rounds[0]
    for match in r0:
        for player in (match.left, match.right):
            if isinstance(player, Participant) and player.name != "BYE":
                player_numbering[player] = counter
                counter += 1

    def label(p: Participant) -> str:
        num = player_numbering.get(p)
        num_str = f"{num:>2}".replace(" ", "\u2007") # 最大2桁で、左をスペースで埋めて表示する
        return f"{num_str}.{p.name.replace('　', ' ')} ({p.kana.replace('　', ' ')}) [{p.team or '-'}]"

    # --- Round0：BYEを詰めた縦配置（各試合 1〜2 行） -------------------------
    line_h         = int(font_size * 1.3)   # ヘッダー各行の行高
    round_label_h  = int(font_size * 1.4)   # その下の「ラウンド名」行の高さ
    extra_pad_px   = 6                      # ちょい余白
    header_h       = line_h * 5 + round_label_h + extra_pad_px # 4+1行分のヘッダ文字列の高さ
    base_y = margin + header_h              # ← 参加者ボックスの開始Yを下げる

    leaf_rows: list[list[float]] = []
    ycur = 0
    for m in r0:
        ys = []
        if is_real(m.left):
            ys.append(base_y + ycur * unit_v); ycur += 1
        if is_real(m.right):
            ys.append(base_y + ycur * unit_v); ycur += 1
        if not ys:  # 保険（通常は発生しない）
            ys.append(base_y + ycur * unit_v); ycur += 1
        leaf_rows.append(ys)
    r0_centers = [(ys[0] if len(ys)==1 else (ys[0]+ys[1])/2) for ys in leaf_rows]

    # --- Round1+ の中心Y ----------------------------------------------------
    round_slot_y: list[list[float]] = [r0_centers]
    prev = r0_centers
    for _ in range(1, R):
        centers = [(prev[i] + prev[i+1]) / 2 for i in range(0, len(prev), 2)]
        round_slot_y.append(centers)
        prev = centers

    # --- X 位置（列）計算 ----------------------------------------------------
    col_x = [margin]
    base_left  = max(int(compact_gap), min_seg)

    alpha = 1.0                 # 見た目の好み（0.8〜0.95目安）
    seg1  = max(int(base_left*alpha), min_seg)  # 選手→子縦線（見える横線）
    seg3  = int(seg1 / 2)                       # 親→次列

    col_x.append(col_x[0] + cell_w + (seg1 * 2) + seg3)
    for _ in range(2, R):
        s_left  = max(int(compact_gap * 0.50), min_seg)
        s_right = max(compact_gap - s_left, min_seg)
        col_x.append(col_x[-1] + s_left + s_right)

    bottom_y = max(y for ys in leaf_rows for y in ys)
    width  = col_x[-1] + margin
    height = int(bottom_y + (cell_h/2) + margin)

    # --- 試合番号（通し） ----------------------------------------------------
    match_no = assign_match_numbers(rounds) if print_match_no == True else None

    # --- SVG出力 -------------------------------------------------------------
    out = []
    out.append(f'<svg xmlns="http://www.w3.org/2000/svg" width="{width}" height="{height}">')
    out.append('<style>')

    # --- 
    out.append(f'text {{ font-size:{font_size}px; font-family:{font_family}; dominant-baseline: middle; }}')
    out.append('.box { fill:#fff; stroke:#666; }')
    out.append('.line { stroke:#666; stroke-width:1; fill:none; }')
    out.append('</style>')

    # --- ヘッダー4行を描画 ---
    y = margin + line_h  # 1行目のY

    # 部門名称
    out.append(f'<text x="0" y="{y}" text-anchor="left" '
               f'style="font-weight:bold; font-size:{font_size+4}px;">{match_name}</text>')
    y += line_h

    # 日付
    out.append(f'<text x="20" y="{y}" text-anchor="left">{match_date}</text>')
    y += line_h

    # 試合会場その１
    out.append(f'<text x="20" y="{y}" text-anchor="left">{match_place1}</text>')
    y += line_h

    # 試合会場その２
    out.append(f'<text x="20" y="{y}" text-anchor="left">{match_place2}</text>')
    y += line_h

    # --- ラウンドラベル（1回戦・2回戦・準決勝・決勝） ---
    label_y = margin + line_h * 4 + int(round_label_h * 0.7) # 4行分のヘッダ文字列を考慮
    
    for r in range(0, R):  # 1回戦はボックス列なので、2回戦以降に表示
        #print(str(r+1) + "回戦")

        # ラウンドの名前を動的に決定
        if r == R - 1:
            label_text = "決勝"
        elif r == R - 2:
            label_text = "準決勝"
        else:
            label_text = f"{r+1}回戦"

        # 交点のX座標
        if r == 0:
            # 1回戦：ボックス→子の縦線(xm_child)
            x_label = (col_x[0] + cell_w) + seg1
        elif r == 1:
            x_label = (col_x[0] + cell_w) + (seg1 * 2)
        else:
            x_label = col_x[r] - (seg1 / 2)

        # ラベル描画
        out.append(f'<text clas="round" x="{x_label}" y="{label_y}" text-anchor="middle" style="font-size:10px;">{label_text}</text>')

    # 1回戦：ボックス（名前のみ）
    x0 = col_x[0]
    for i, m in enumerate(r0):
        ys = leaf_rows[i]
        k = 0
        if is_real(m.left):
            y = ys[k]; k += 1
            out.append(f'<rect class="box" x="{x0}" y="{y-cell_h/2}" width="{cell_w}" height="{cell_h}" rx="6" ry="6"/>')
            out.append(f'<text x="{x0+8}" y="{y}">{esc(label(m.left))}</text>')
        if is_real(m.right):
            y = ys[k]; k += 1
            out.append(f'<rect class="box" x="{x0}" y="{y-cell_h/2}" width="{cell_w}" height="{cell_h}" rx="6" ry="6"/>')
            out.append(f'<text x="{x0+8}" y="{y}">{esc(label(m.right))}</text>')

    # 1回戦→2回戦：線 + 1回戦番号 + 2回戦番号
    if R >= 2:
        x_prev_right = x0 + cell_w
        xm_child  = x_prev_right + seg1
        xm_parent = xm_child   + seg1
        x_parent_left = col_x[1]

        for i in range(len(rounds[1])):
            ysL = leaf_rows[2*i]
            ysR = leaf_rows[2*i+1]

            # 左子
            if len(ysL) == 2:
                out.append(f'<path class="line" d="M{x_prev_right},{ysL[0]} H{xm_child}"/>')
                out.append(f'<path class="line" d="M{x_prev_right},{ysL[1]} H{xm_child}"/>')
                out.append(f'<path class="line" d="M{xm_child},{ysL[0]} V{ysL[1]}"/>')
                cL = (ysL[0] + ysL[1]) / 2
            else:
                out.append(f'<path class="line" d="M{x_prev_right},{ysL[0]} H{xm_child}"/>')
                cL = ysL[0]
            # 1回戦番号（BYE含む試合は採番されていない）
            nL = match_no.get((0, 2*i)) if match_no != None else None
            if nL is not None:
                out.append(f'<text x="{xm_child-8}" y="{cL}" text-anchor="end" fill="#333">{nL}</text>')
            out.append(f'<path class="line" d="M{xm_child},{cL} H{xm_parent}"/>')

            # 右子
            if len(ysR) == 2:
                out.append(f'<path class="line" d="M{x_prev_right},{ysR[0]} H{xm_child}"/>')
                out.append(f'<path class="line" d="M{x_prev_right},{ysR[1]} H{xm_child}"/>')
                out.append(f'<path class="line" d="M{xm_child},{ysR[0]} V{ysR[1]}"/>')
                cR = (ysR[0] + ysR[1]) / 2
            else:
                out.append(f'<path class="line" d="M{x_prev_right},{ysR[0]} H{xm_child}"/>')
                cR = ysR[0]
            nR = match_no.get((0, 2*i+1)) if match_no != None else None
            if nR is not None:
                out.append(f'<text x="{xm_child-8}" y="{cR}" text-anchor="end" fill="#333">{nR}</text>')
            out.append(f'<path class="line" d="M{xm_child},{cR} H{xm_parent}"/>')

            # 親（2回戦）
            cy = round_slot_y[1][i]
            out.append(f'<path class="line" d="M{xm_parent},{cL} V{cR}"/>')
            out.append(f'<path class="line" d="M{xm_parent},{cy} H{x_parent_left}"/>')
            nP = match_no.get((1, i)) if match_no != None else None
            if nP is not None:
                out.append(f'<text x="{xm_parent-10}" y="{cy}" text-anchor="end" fill="#333">{nP}</text>')

    # 3回戦以降
    for r in range(2, R):
        prev_c = round_slot_y[r-1]
        c      = round_slot_y[r]
        x_prev_right = col_x[r-1]
        s_left  = col_x[r] - col_x[r-1] - seg3
        s_right = seg3
        xm_parent    = x_prev_right + s_left
        x_parent_left= x_prev_right + s_left + s_right

        for i in range(len(c)):
            cL = prev_c[2*i]; cR = prev_c[2*i+1]; cy = c[i]
            out.append(f'<path class="line" d="M{x_prev_right},{cL} H{xm_parent}"/>')
            out.append(f'<path class="line" d="M{x_prev_right},{cR} H{xm_parent}"/>')
            out.append(f'<path class="line" d="M{xm_parent},{cL} V{cR}"/>')
            out.append(f'<path class="line" d="M{xm_parent},{cy} H{x_parent_left}"/>')

            n = match_no.get((r, i)) if match_no != None else None
            if n is not None:
                out.append(f'<text x="{xm_parent-8}" y="{cy}" text-anchor="end" fill="#333">{n}</text>')

    out.append('</svg>')
    with open(path, "w", encoding="utf-8") as f:
        f.write("\n".join(out))


# --- エクセル出力 ---
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Side
from typing import List

def save_bracket_xlsx(
    rounds: List[Round],
    path="bracket.xlsx",
    init_workbook = False,
    summary_name = "xxxxxxxxの部", # シート名
    match_name = "xxxxxxxxの部", # 出力する部門名
    match_date = "2099.12.31",
    match_place1 = "第一試合場",
    match_place2 = "第二試合場",
):
    # ─────────────────────────────
    #  Excel 初期化
    # ─────────────────────────────
    if init_workbook:
        # 既存のファイルを削除し、新しいエクセルファイルを作成する
        try:
            os.remove(path)
            print(f"[{path}] を削除しました。")
        except FileNotFoundError:
            pass
        wb = Workbook()

        # 新しいシート{summary_name}を作成し、それ以外のデフォルトのシートを削除する
        ws = wb.create_sheet(summary_name)
        for sheet_name in wb.sheetnames:
            if sheet_name != summary_name:
                wb.remove(wb[sheet_name])
        wb.save(path)

    else:
        # 既存のエクセルファイルを開く
        wb = load_workbook(path)

        # 既存のシート{summary_name}を削除し、新しいシート{summary_name}を作成する
        if summary_name in  wb.sheetnames:
            wb.remove(wb[summary_name])
        ws = wb.create_sheet(summary_name)
        wb.save(path)

    ws.title = summary_name
    ws.column_dimensions["A"].width = ws.column_dimensions["A"].width / 2 # 選手番号の列: デフォルトの1/2
    ws.column_dimensions["B"].width = ws.column_dimensions["B"].width * 1.5 # 選手名称の列: デフォルトの2倍

    thin = Side(style="thin")

    # ─────────────────────────────
    #  helpers
    # ─────────────────────────────
    def is_real(x: Slot) -> bool:
        return isinstance(x, Participant) and x.name != "BYE"

    def assign_match_numbers(rounds: List[Round]) -> dict[tuple[int, int], int]:
        """BYEを除く全ラウンドの実施試合に通し番号を振る。"""
        n = 1
        mapping = {}
        for r, rnd in enumerate(rounds):
            for i, m in enumerate(rnd):
                # 左右のどちらも BYE でない（＝実際に行われる試合）
                left_is_bye  = isinstance(m.left, Participant)  and m.left.name  == "BYE"
                right_is_bye = isinstance(m.right, Participant) and m.right.name == "BYE"
                if not (left_is_bye or right_is_bye):
                    mapping[(r, i)] = n
                    n += 1
        return mapping

    match_no = assign_match_numbers(rounds)

    R = len(rounds)

    # ─────────────────────────────
    #  部門情報
    # ─────────────────────────────
    row = 1
    ws.cell(row, 1).value = match_name
    row += 1
    ws.cell(row, 2).value = match_date
    row += 1
    ws.cell(row, 2).value = match_place1
    row += 1
    ws.cell(row, 2).value = match_place2

    # ─────────────────────────────
    #  "1回戦", ... "準決勝", "決勝" のラベル出力
    # ─────────────────────────────
    row += 1
    for r in range(0, R):
        if r == R - 1:
            ws.cell(row, r + 4).value = "決勝"
        elif r == R - 2:
            ws.cell(row, r + 4).value = "準決勝"
        else:
            ws.cell(row, r + 4).value = f"{r + 1}回戦"
        ws.cell(row, r + 4).alignment = Alignment(horizontal="right")

    # ─────────────────────────────
    #  選手情報
    # ─────────────────────────────
    row_player_1st = row + 1 # 最初の選手の行を保存しておく

    # 選手番号のマップを作成
    player_numbering: dict[Participant, int] = {}
    player_counter = 1
    r0 = rounds[0]
    for match in r0:
        for player in (match.left, match.right):
            if is_real(player):
                player_numbering[player] = player_counter
                player_counter += 1

    def output_player(player):
        player_no = player_numbering.get(player)
        row = row_player_1st + (player_no - 1) * 2
        ws.cell(row, 1).value = player_no
        ws.cell(row, 2).value = player.name
        ws.cell(row, 3).value = f"[{player.team}]"
        ws.cell(row + 1, 2).value = f"({player.kana})"

        # セル結合
        ws.merge_cells(start_row=row, start_column=1, end_row=row+1, end_column=1)
        ws.merge_cells(start_row=row, start_column=3, end_row=row+1, end_column=3)

        # 文字位置調整
        ws.cell(row, 1).alignment = Alignment(horizontal="right", vertical="center")
        ws.cell(row, 2).alignment = Alignment(horizontal="left", vertical="bottom")
        ws.cell(row+1, 2).alignment = Alignment(horizontal="left", vertical="top")
        ws.cell(row, 3).alignment = Alignment(horizontal="left", vertical="center")

        # 縦幅を拡張し、選手と選手の間にスペースを設ける（縦幅のデフォルトは約15らしい）
        ws.row_dimensions[row].height = 15 * 1.25
        ws.row_dimensions[row+1].height = 15 * 1.25

    for match in r0:
        if is_real(match.left):
            output_player(match.left)

        if is_real(match.right):
            output_player(match.right)

    # ─────────────────────────────
    #  トーナメント線、試合番号
    # ─────────────────────────────
    match_2_row = {} # KEY:match、VALUE:match勝者の水平線の直上行番号

    for rnd_idx, rnd in enumerate(rounds):
        rnd_no = rnd_idx + 1
        rnd_col = rnd_no + 3
        
        # 1回戦
        if rnd_no == 1:
            for match_idx, match in enumerate(rnd):
                if is_real(match.left):
                    # 選手は2行で構成されるが、その1行目の底辺に水平線を引く
                    row = row_player_1st + (player_numbering.get(match.left) - 1) * 2
                    ws.cell(row, rnd_col).border = Border(bottom=thin)

                if is_real(match.right):
                    # 選手は2行で構成されるが、その2行目の上辺に水平線を引く
                    row = row_player_1st + (player_numbering.get(match.right) - 1) * 2
                    ws.cell(row + 1, rnd_col).border = Border(top=thin)
                
                if is_real(match.left) and is_real(match.right):
                    # 二人の選手の水平線に囲まれたセルを結合し、右辺に垂直線を引く
                    row_left = row_player_1st + (player_numbering.get(match.left) - 1) * 2
                    row_right = row_player_1st + (player_numbering.get(match.right) - 1) * 2
                    ws.merge_cells(start_row=row_left+1, start_column=rnd_col, end_row=row_right, end_column=rnd_col) # 結合
                    for row in range(row_left+1, row_right + 1):
                        ws.cell(row, rnd_col).border = Border(right=thin) # 垂直線
                    
                    # 試合番号
                    ws.cell(row_left + 1, rnd_col).value = match_no.get((rnd_idx, match_idx))
                    ws.cell(row_left + 1, rnd_col).alignment = Alignment(horizontal="right", vertical="center")
                
                # match勝者の水平線の直上行番号を記録する
                if is_real(match.left) and is_real(match.right):
                    match_2_row[(rnd_idx, match_idx)] = row_player_1st + (player_numbering.get(match.left) - 1) * 2 + 1
                elif is_real(match.left):
                    match_2_row[(rnd_idx, match_idx)] = row_player_1st + (player_numbering.get(match.left) - 1) * 2
                elif is_real(match.right):
                    match_2_row[(rnd_idx, match_idx)] = row_player_1st + (player_numbering.get(match.right) - 1) * 2

        # 2回戦～決勝
        else:
            for match_idx, match in enumerate(rnd):
                # 選手1の行の底辺に水平線を引く
                row_left = match_2_row[(rnd_idx - 1, match_idx * 2)]
                ws.cell(row_left, rnd_col).border = Border(bottom=thin)

                # 選手2の行+1の上辺に水平線を引く
                row_right = match_2_row[(rnd_idx - 1, match_idx * 2 + 1)]
                ws.cell(row_right + 1, rnd_col).border = Border(top=thin)

                # 二人の選手の水平線に囲まれたセルを結合し、右辺に垂直線を引く
                ws.merge_cells(start_row=row_left+1, start_column=rnd_col, end_row=row_right, end_column=rnd_col) # 結合
                for row in range(row_left+1, row_right + 1):
                    ws.cell(row, rnd_col).border = Border(right=thin) # 垂直線

                # 試合番号
                ws.cell(row_left + 1, rnd_col).value = match_no.get((rnd_idx, match_idx))
                ws.cell(row_left + 1, rnd_col).alignment = Alignment(horizontal="right", vertical="center")

                # match勝者の水平線の直上行番号を記録する
                match_2_row[(rnd_idx, match_idx)] = (row_left + row_right) // 2

    # 決勝戦の垂直線の中心から右に水平線を引く
    rnd_col = 3 + len(rounds) + 1
    ws.cell(match_2_row[len(rounds) - 1, 0], rnd_col).border = Border(bottom=thin)

    # ファイルを保存する
    wb.save(path)


# --------------------
# サンプル実行
# --------------------
if __name__ == "__main__":
    sample = [
        Participant("伊藤　創", "いとう　はじめ", "中央会"),
        Participant("荒井　裕幸", "あらい　ひろゆき", "中央会"),
        Participant("柳田　哲朗", "やなぎだ　てつお", "中央会"),
        Participant("齋藤　直樹", "さいとう　なおき", "中央会"),
        Participant("石黒　健司", "いしぐろ　けんじ", "中央会"),
        Participant("塩幡　勝典", "しおはた　かつのり", "中央会"),
        Participant("杉野　仁司", "すぎの　ひとし", "中央会"),
        Participant("福田　知広", "ふくだ　かずひろ", "中央会"),
        Participant("北村 和也", "きたむら　かずや", "大町"),
        Participant("富宇加　健", "とみうか　けん", "聖武会"),
        Participant("中嶋　昭弘", "なかじま　あきひろ", "深大寺"),
        Participant("磯野　渉", "いその　わたる", "深大寺"),
        Participant("松本　義弘", "まつもと　よしひろ", "深大寺"),
        Participant("國分　崇生", "こくぶん　たかお", "電通大"),
        #Participant("RX伊藤　創", "いとう　はじめ", "中央会"),
        #Participant("RX荒井　裕幸", "あらい　ひろゆき", "中央会"),
        #Participant("RX柳田　哲朗", "やなぎだ　てつお", "中央会"),
        #Participant("RX齋藤　直樹", "さいとう　なおき", "中央会"),
        #Participant("RX石黒　健司", "いしぐろ　けんじ", "中央会"),
        #Participant("RX塩幡　勝典", "しおはた　かつのり", "中央会"),
        #Participant("RX杉野　仁司", "すぎの　ひとし", "中央会"),
        #Participant("RX福田　知広", "ふくだ　かずひろ", "中央会"),
        #Participant("RX北村 和也", "きたむら　かずや", "大町"),
        #Participant("RX富宇加　健", "とみうか　けん", "聖武会"),
        #Participant("RX中嶋　昭弘", "なかじま　あきひろ", "深大寺"),
        #Participant("RX磯野　渉", "いその　わたる", "深大寺"),
        #Participant("RX松本　義弘", "まつもと　よしひろ", "深大寺"),
        #Participant("RX國分　崇生", "こくぶん　たかお", "電通大"),
    ]
    
    rounds = build_full_bracket(sample, seed=1)
    
    print_bracket(rounds)
    save_bracket_svg(rounds, "result.svg")
    save_bracket_xlsx(rounds, "result.xlsx", match_name="一般六･七段の部", match_date="2099.12.31", match_place1="第一試合場(選手番号1～7)", match_place2="第二試合場(選手番号8～14)")
