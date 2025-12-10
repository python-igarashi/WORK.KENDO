from __future__ import annotations
from dataclasses import dataclass
from collections import defaultdict
from typing import List, Tuple, Optional, Union, Set
import math, random, os

# --------------------
# 基本データ構造
# --------------------
@dataclass(frozen=True)
class Participant:
	name: str
	kana: str # ふりがな
	groupname: Optional[str]  # 団体名（なしなら None）

@dataclass(frozen=True)
class Ref:
	"""前ラウンドの '試合番号' を参照（その勝者がここに入る）"""
	match_index: int  # 0-based within that round

Slot = Union[Participant, Ref]

@dataclass
class Match:
	left: Slot
	right: Slot
	winner: Optional[Participant] = None  # BYE確定などで分かる場合のみ埋める

Round = List[Match]

# --------------------
# ユーティリティ
# --------------------
def next_power_of_two(n: int) -> int:
	return 1 if n <= 1 else 2 ** math.ceil(math.log2(n))

def _is_bye(p) -> bool:
	return (p is None) or (getattr(p, "name", None) == "BYE")

def _groupname(p) -> str | None:
	if _is_bye(p):
		return None
	return getattr(p, "groupname", None) or None

def bye_positions_balanced_halves(M: int, B: int) -> list[int]:
	"""
	総試合数 M を 4 クォーターに分けつつ、
	まず Top(=Q1+Q2) / Bottom(=Q3+Q4) で BYE を均等配分し、
	その後 各半分の2クォーターへ均等配分して配置位置(0..M-1)を返す。
	"""
	if B <= 0:
		return []

	# 1) クォーターサイズ（Mを4等分、端数は上から配分）
	base = M // 4
	rem  = M % 4
	q_sizes = [(base + (1 if i < rem else 0)) for i in range(4)]  # [Q1,Q2,Q3,Q4]

	# 2) 半分(TOP/BOTTOM)でBYE数を均等化
	#   half_byes[0]=Top, half_byes[1]=Bottom
	half_byes = [B // 2, B // 2]
	if B % 2 == 1:
		# 端数は、クォーター合計の大きい方(Top or Bottom)に付与
		top_size = q_sizes[0] + q_sizes[1]
		bot_size = q_sizes[2] + q_sizes[3]
		if top_size >= bot_size:
			half_byes[0] += 1
		else:
			half_byes[1] += 1

	# 3) 各半分の2クォーターへ均等配分（端数は“広い方”のクォーターへ）
	q_byes = [0, 0, 0, 0]  # [Q1,Q2,Q3,Q4]

	def split_into_two(total_byes: int, size_a: int, size_b: int) -> tuple[int, int]:
		a = total_byes // 2
		b = total_byes - a
		# 端数配分の調整：広い方に多めを寄せる
		if size_a < size_b and a > b:
			a, b = b, a
		if size_b < size_a and b > a:
			a, b = b, a
		return a, b

	q_byes[0], q_byes[1] = split_into_two(half_byes[0], q_sizes[0], q_sizes[1])  # Top -> Q1,Q2
	q_byes[2], q_byes[3] = split_into_two(half_byes[1], q_sizes[2], q_sizes[3])  # Bottom -> Q3,Q4

	# 4) 各クォーター内で等間隔の試合に配置 → グローバルindexへ変換
	res = []
	offset = 0
	for q in range(4):
		size = q_sizes[q]
		b	= q_byes[q]
		if size == 0 or b == 0:
			# 何も置かない
			pass
		else:
			# 等間隔配置（重複時は近傍にずらす）
			local = []
			used = set()
			for i in range(b):
				idx = round((i + 0.5) * size / b) - 1
				idx = max(0, min(size - 1, idx))
				if idx in used:
					# 近い空きを探す
					found = None
					for d in range(1, size):
						for cand in (idx - d, idx + d):
							if 0 <= cand < size and cand not in used:
								found = cand
								break
						if found is not None:
							break
					idx = found if found is not None else idx
				used.add(idx)
				local.append(idx)
			res.extend([offset + x for x in sorted(local)])
		offset += size

	return sorted(res)

# --------------------
# “同団体回避”を意識した並べ方（バケット→ラウンドロビン）
# --------------------
def diversified_order(participants: List[Participant], seed: Optional[int]) -> List[Participant]:
	"""チーム分散しつつ、並びの規則性を弱める"""
	import random
	rnd = random.Random(seed)

	# バケット化 & 各バケット内はシャッフル
	buckets = defaultdict(list)
	for p in participants:
		buckets[p.groupname].append(p)
	for k in buckets:
		rnd.shuffle(buckets[k])

	# ① チームを「人数降順」でグルーピングし、同サイズ内はランダム
	by_size = defaultdict(list)
	for groupname, members in buckets.items():
		by_size[len(members)].append(groupname)

	sizes = sorted(by_size.keys(), reverse=True)
	bucket_keys: List[Optional[str]] = []
	for s in sizes:
		ks = by_size[s][:]
		rnd.shuffle(ks)  # ← 同規模チームの順序をランダム化
		bucket_keys.extend(ks)

	# ② 先頭の開始オフセットもランダムに回転
	if bucket_keys:
		off = rnd.randrange(len(bucket_keys))
		bucket_keys = bucket_keys[off:] + bucket_keys[:off]

	# ③ 取り出し周回ごとに左右反転（serpentine）して規則性を崩す
	order: List[Participant] = []
	forward = True
	while any(buckets[k] for k in bucket_keys):
		keys = bucket_keys if forward else list(reversed(bucket_keys))
		for k in keys:
			if buckets[k]:
				order.append(buckets[k].pop())
		forward = not forward

	# --- 上下偏り軽減のため、最終的な並びを上下で交互ミックス ---
	mid = len(order) // 2
	top = order[:mid]
	bottom = order[mid:]
	mixed = []
	for i in range(max(len(top), len(bottom))):
		if i < len(top):
			mixed.append(top[i])
		if i < len(bottom):
			mixed.append(bottom[i])

	return mixed

# --------------------
# 1回戦：四分割均等BYE割当（BYE vs BYE なし）
# --------------------
Pair = Tuple[Participant, Participant]

from collections import Counter
from typing import Optional

from collections import Counter
from typing import Optional

def _assign_pairs_with_bye_and_lookahead(
	order: list["Participant"],
	M: int,
	bye_pos: set[int],
	lookahead: int,
	rnd,
) -> list[tuple["Participant", "Participant"]]:
	"""
	1回戦のペアを構築する（half分散対応版）。

	目的：
	- v1（初戦の同団体対戦）は可能な限り0（ハード最優先）
	  -> b選びは同団体を窓→全体で探索して回避
	  -> a選びも「相手(別団体)が残る」候補を優先して詰みを回避
	- half（上半分/下半分）への団体偏りを抑える（ソフト制約）
	  -> 団体gの総人数 total[g] に対し、各halfへ ceil(total[g]/2) を超えて置かないよう努力
	  -> どうしても無理なら妥協する
	- v2/v3（親ペア内のクロス）は、2試合目のa側で used を避けるなど、できる範囲で抑える（ソフト）
	"""

	bye = Participant("BYE", "BYE", "-")

	# order読み取り位置
	idx = 0

	# resultは最初にBYEで埋める（後で _set_match で確定）
	result: list[tuple["Participant", "Participant"]] = [(bye, bye) for _ in range(M)]

	# --- half分散（上半分/下半分）のソフト制約用 ---
	total_by_group = Counter()
	for p in order:
		g = _groupname(p)
		if g:
			total_by_group[g] += 1

	placed_half = {g: [0, 0] for g in total_by_group.keys()}  # [upper, lower]

	def half_of_match(mi: int) -> int:
		# 0=上半分, 1=下半分
		return 0 if mi < (M // 2) else 1

	def half_ok(g: Optional[str], half: int) -> bool:
		# BYE等は気にしない
		if not g:
			return True
		# ceil(total/2) をそのhalfへのソフト上限とする
		limit = (total_by_group[g] + 1) // 2
		return placed_half[g][half] < limit

	def count_place(p: "Participant", mi: int) -> None:
		g = _groupname(p)
		if not g:
			return
		h = half_of_match(mi)
		placed_half[g][h] += 1

	def set_match(mi: int, a: "Participant", b: "Participant") -> None:
		result[mi] = (a, b)
		if not _is_bye(a):
			count_place(a, mi)
		if not _is_bye(b):
			count_place(b, mi)

	# --- orderの要素を idx へ引っ張る操作 ---
	def pull_to_front(j: int) -> "Participant":
		nonlocal idx
		p = order[j]
		for k in range(j, idx, -1):
			order[k] = order[k - 1]
		order[idx] = p
		idx += 1
		return p

	def pop_next() -> "Participant":
		nonlocal idx
		p = order[idx]
		idx += 1
		return p

	def remaining_group_counter() -> Counter:
		c = Counter()
		for p in order[idx:]:
			g = _groupname(p)
			if g:
				c[g] += 1
		return c

	# --- a/b選択ロジック ---

	def pick_a_feasible(target_half: int) -> "Participant":
		"""
		a（先手）：
		- “別団体の相手が残る” 候補を優先（詰み回避）
		- half_ok をできれば守る（ソフト）
		"""
		nonlocal idx
		if idx >= len(order):
			return bye

		cnt = remaining_group_counter()

		def feasible_group(g: Optional[str]) -> bool:
			if not g:
				return True
			# g以外が残っていればOK
			return (sum(cnt.values()) - cnt.get(g, 0)) > 0

		def ok(j: int) -> bool:
			g = _groupname(order[j])
			return feasible_group(g) and half_ok(g, target_half)

		w_end = min(idx + lookahead, len(order))

		# 1) 窓：feasible かつ half_ok
		cands = [j for j in range(idx, w_end) if ok(j)]
		if cands:
			return pull_to_front(rnd.choice(cands))

		# 2) 窓：feasible（half妥協）
		cands = [j for j in range(idx, w_end) if feasible_group(_groupname(order[j]))]
		if cands:
			return pull_to_front(rnd.choice(cands))

		# 3) 全体：feasible かつ half_ok
		cands = [j for j in range(idx, len(order)) if ok(j)]
		if cands:
			return pull_to_front(rnd.choice(cands))

		# 4) 全体：feasible（half妥協）
		cands = [j for j in range(idx, len(order)) if feasible_group(_groupname(order[j]))]
		if cands:
			return pull_to_front(rnd.choice(cands))

		return pop_next()

	def pick_b_avoid_group(hard_avoid: Optional[str], target_half: int) -> "Participant":
		"""
		b（相手）：
		- hard_avoid（=aの団体）を回避（v1ハード）
		- half_ok をできれば守る（ソフト）
		"""
		nonlocal idx
		if idx >= len(order):
			return bye

		def ok(j: int) -> bool:
			g = _groupname(order[j])
			if hard_avoid and g == hard_avoid:
				return False
			return half_ok(g, target_half)

		w_end = min(idx + lookahead, len(order))

		# 1) 窓：hard回避 & half_ok
		cands = [j for j in range(idx, w_end) if ok(j)]
		if cands:
			return pull_to_front(rnd.choice(cands))

		# 2) 窓：hard回避（half妥協）
		if hard_avoid:
			cands = [j for j in range(idx, w_end) if _groupname(order[j]) != hard_avoid]
			if cands:
				return pull_to_front(rnd.choice(cands))
		else:
			# hard制約が無いときでも、窓でhalf_okを優先したい
			cands = [j for j in range(idx, w_end) if half_ok(_groupname(order[j]), target_half)]
			if cands:
				return pull_to_front(rnd.choice(cands))

		# 3) 全体：hard回避 & half_ok
		cands = [j for j in range(idx, len(order)) if ok(j)]
		if cands:
			return pull_to_front(rnd.choice(cands))

		# 4) 全体：hard回避（half妥協）
		if hard_avoid:
			cands = [j for j in range(idx, len(order)) if _groupname(order[j]) != hard_avoid]
			if cands:
				return pull_to_front(rnd.choice(cands))
		else:
			cands = [j for j in range(idx, len(order)) if half_ok(_groupname(order[j]), target_half)]
			if cands:
				return pull_to_front(rnd.choice(cands))

		return pop_next()

	def pick_soft_avoid(avoid: set[str], target_half: int) -> "Participant":
		"""
		ソフト選択：
		- avoid（できれば避けたい団体集合）を避ける（v2/v3低減）
		- half_ok もできれば守る
		"""
		nonlocal idx
		if idx >= len(order):
			return bye

		def ok(j: int) -> bool:
			g = _groupname(order[j])
			return half_ok(g, target_half) and ((g is None) or (g not in avoid))

		w_end = min(idx + lookahead, len(order))

		# 1) 窓：avoid回避 & half_ok
		cands = [j for j in range(idx, w_end) if ok(j)]
		if cands:
			return pull_to_front(rnd.choice(cands))

		# 2) 窓：half_ok（avoid妥協）
		cands = [j for j in range(idx, w_end) if half_ok(_groupname(order[j]), target_half)]
		if cands:
			return pull_to_front(rnd.choice(cands))

		# 3) 全体：avoid回避 & half_ok
		cands = [j for j in range(idx, len(order)) if ok(j)]
		if cands:
			return pull_to_front(rnd.choice(cands))

		# 4) 全体：half_ok（avoid妥協）
		cands = [j for j in range(idx, len(order)) if half_ok(_groupname(order[j]), target_half)]
		if cands:
			return pull_to_front(rnd.choice(cands))

		return pop_next()

	# --- 親ペア（2試合）単位で作る ---
	parent_starts = list(range(0, M, 2))
	rnd.shuffle(parent_starts)

	for i0 in parent_starts:
		i1 = i0 + 1
		if i1 >= M:
			i1 = i0

		h0 = half_of_match(i0)
		h1 = half_of_match(i1)

		is_bye0 = (i0 in bye_pos)
		is_bye1 = (i1 in bye_pos)

		# 両方BYE試合枠（= (P,BYE) と (Q,BYE)）
		if is_bye0 and is_bye1:
			a = pick_a_feasible(h0)
			b = pick_b_avoid_group(_groupname(a), h1)  # v1を避けつつhalfも意識
			set_match(i0, a, bye)
			set_match(i1, b, bye)
			continue

		# BYEなし親（=2試合とも実試合）
		if (not is_bye0) and (not is_bye1):
			# 1試合目（v1ハード）
			a1 = pick_a_feasible(h0)
			b1 = pick_b_avoid_group(_groupname(a1), h0)
			set_match(i0, a1, b1)

			used = {g for g in (_groupname(a1), _groupname(b1)) if g}

			# 2試合目：a2 は used をできれば避ける（ソフト）
			a2 = pick_soft_avoid(used, h1)
			b2 = pick_b_avoid_group(_groupname(a2), h1)  # v1ハード
			set_match(i1, a2, b2)
			continue

		# 片BYE親：通常試合→BYE側を「通常試合の団体」を避けつつ置く
		if is_bye0 and (not is_bye1):
			# i1が通常試合
			aN = pick_a_feasible(h1)
			bN = pick_b_avoid_group(_groupname(aN), h1)
			set_match(i1, aN, bN)

			avoid = {g for g in (_groupname(aN), _groupname(bN)) if g}
			aB = pick_soft_avoid(avoid, h0)
			set_match(i0, aB, bye)
			continue

		if is_bye1 and (not is_bye0):
			# i0が通常試合
			aN = pick_a_feasible(h0)
			bN = pick_b_avoid_group(_groupname(aN), h0)
			set_match(i0, aN, bN)

			avoid = {g for g in (_groupname(aN), _groupname(bN)) if g}
			aB = pick_soft_avoid(avoid, h1)
			set_match(i1, aB, bye)
			continue

	return result

def _future_round_penalty(
    pairs: List[Pair],
    base: int = 10**8,
    decay: int = 50,
    exclude_final: bool = True
) -> int:
    """
    “実際の対戦可能性” に対するペナルティ
      - 1回戦の葉を 0..T-1 に番号付け
      - 各団体について、所属選手の葉インデックスの全ペアを調べる
      - その2人が最短で当たり得るラウンド r を計算し、
        base / decay^r の重みでペナルティ加算
      - exclude_final=True の場合、決勝で初めて当たるペアは無視
    """

    # 1回戦の葉インデックス割り当て
    # Match i の left -> idx = 2*i, right -> idx = 2*i+1
    slots: List[Slot] = []
    for m in pairs:
        slots.append(m[0])
        slots.append(m[1])

    T = len(slots)               # 葉の総数（2のべきのはず）
    if T == 0:
        return 0

    max_round = int(math.log2(T)) - 1  # 0=1回戦, 1=2回戦,..., max_round=決勝

    def team_of(p) -> Optional[str]:
        if p is None:
            return None
        if getattr(p, "name", None) == "BYE":
            return None
        t = _groupname(p)
        return t if t else None

    # 団体ごとに葉インデックスを集計
    team_to_indices: dict[str, List[int]] = defaultdict(list)
    for idx, p in enumerate(slots):
        t = team_of(p)
        if t:
            team_to_indices[t].append(idx)

    total_penalty = 0

    # 各団体について、所属選手のペアごとに最短対戦ラウンドを評価
    for t, indices in team_to_indices.items():
        n = len(indices)
        if n <= 1:
            continue  # 1人では同団体対戦は起こらない

        for a_idx in range(n):
            i = indices[a_idx]
            for b_idx in range(a_idx + 1, n):
                j = indices[b_idx]

                x = i ^ j
                if x == 0:
                    # 同じスロットにいることは通常ないが、念のため
                    continue

                r = int(math.log2(x))  # 0=1回戦, 1=2回戦,...

                # 決勝のみペナルティ除外したい場合
                if exclude_final and r == max_round:
                    continue

                weight = base // (decay ** r)
                if weight <= 0:
                    weight = 1
                total_penalty += weight

    return total_penalty

def _half_distribution_penalty(pairs: List[Pair], base_half: int = 10**7) -> int:
    """
    各団体が「上の山」と「下の山」にどれくらい均等に分かれているかのペナルティ。
      - 試合 index < M/2 を上の山
      - index >= M/2 を下の山と見なす
      - 各団体ごとに abs(top - bottom)^2 * base_half を加算
      - その団体の参加人数が1人だけの場合はペナルティ0（分けようがないので）
    """
    M = len(pairs)
    half_matches = M // 2

    # groupname -> [top_count, bottom_count]
    counts: dict[str, list[int]] = defaultdict(lambda: [0, 0])

    for i, (a, b) in enumerate(pairs):
        half = 0 if i < half_matches else 1  # 0: 上の山, 1: 下の山
        for p in (a, b):
            # BYEは無視
            if getattr(p, "name", None) == "BYE":
                continue
            t = _groupname(p)
            if not t:
                continue
            counts[t][half] += 1

    penalty = 0
    for t, (top, bottom) in counts.items():
        total = top + bottom
        if total <= 1:
            # 1人だけの団体はどうしようもないのでペナルティなし
            continue
        diff = abs(top - bottom)
        penalty += (diff ** 2) * base_half

    return penalty

def _conflicts(pairs: List[Pair]) -> int:
	"""
	衝突スコアを返す。
	- 1回戦の同一団体対戦: +1
	- 2回戦の同一団体対戦: +1   ※(BYE,P) vs (BYE,Q) で, P と Q が同一団体: +3（2回戦で直接同一団体）
	"""

	v1 = 0 # 第一優先チェック
	v2 = 0 # 第二優先チェック
	v3 = 0 # 第三優先チェック

	# 1) 1回戦の同一団体 → 可能な限り避ける
	for a, b in pairs:
		if not _is_bye(a) and not _is_bye(b):
			ta = _groupname(a)
			tb = _groupname(b)
			if ta and tb and ta == tb:
				v1 += 1

	# 2) 2回戦が初戦の衝突（親ノード=2試合ごと）
	for i in range(0, len(pairs), 2):
		(aL, aR) = pairs[i]
		(bL, bR) = pairs[i+1]

		a_real = [x for x in (aL, aR) if not _is_bye(x)]
		b_real = [x for x in (bL, bR) if not _is_bye(x)]

		a_has_bye = (len(a_real) == 1)
		b_has_bye = (len(b_real) == 1)

		# (BYE,P) vs (BYE,Q): P と Q は2回戦で直接対戦 → 可能な限り避ける
		if a_has_bye and b_has_bye and a_real and b_real:
			tP, tQ = _groupname(a_real[0]), _groupname(b_real[0])
			if tP and tQ and tP == tQ:
				v1 += 1

		# (BYE,P) vs (Q1,Q2): P は Q1/Q2 のどちらとも同一団体NG → 軽く避ける
		# (P1,P2) vs (Q1,Q2): P1/P2 は Q1/Q2 のどちらとも同一団体NG → 軽く避けるが、ある程度やむを得ない事とする
		else:
			for P in a_real:
				for Q in b_real:
					tP, tQ = _groupname(P), _groupname(Q)
					if tP and tQ and tP == tQ:
						if a_has_bye or b_has_bye:
							v2 += 1
						else:
							v3 += 1
	
	# 3) 3回戦のチェック（親ノード=4試合ごと）
	for i in range(0, len(pairs), 4):
		end = min(i + 4, len(pairs))
		if end - i < 4:
			continue  # 4試合揃っていない端数ブロックは対象外

		bye_winners = []
		for ii in range(i, end):
			(aL, aR) = pairs[ii]
			a_real = [x for x in (aL, aR) if not _is_bye(x)]
			if len(a_real) == 1:
				bye_winners.append(a_real[0])

		# BYE勝ち上がりが多いブロックだけチェック
		if 2 <= len(bye_winners) <= 4:   # 必要なら 3<=...<=4 に
			for x in range(len(bye_winners) - 1):
				for y in range(x + 1, len(bye_winners)):
					tP, tQ = _groupname(bye_winners[x]), _groupname(bye_winners[y])
					if tP and tQ and tP == tQ:
						v3 += 1

	return (v1, v2, v3)

def make_first_round_pairs_quarter_even(
	participants: List[Participant],
	seed: Optional[int] = None,
	restarts: int = 1000,	 # 複数リスタート
	lookahead: int = None, #6,	  # 先読み窓幅（団体衝突を回避）
	max_local_iters: int = 2000,
) -> List[Pair]:
	"""
	四分割BYE配分を維持しながら、同団体の初戦を強く回避。
	- diversified_order のランダム性は活かす
	- 先読み割付 + 強化ローカル最適化
	- 最良案（同団体カード数最小）を採用
	"""
	N = len(participants)
	T = next_power_of_two(N)
	M = T // 2
	B = T - N
	bye_pos = set(bye_positions_balanced_halves(M, B))

	# 参加人数(N)に応じて団体衝突回避用変数を大きくする
	if lookahead == None:
		if N <= 32:
			lookahead = 8
		elif N <= 64:
			lookahead = 10
		else:
			lookahead = 16
	print(f"N={N}, lookahead={lookahead}")

	import random
	base_rnd = random.Random(seed)
	best_pairs = None
	best_c1 = 10**9
	best_c2 = 10**9
	best_c3 = 10**9
	best_c4 = 10**11
	best_c5 = 10**11

	best_score = None
	for r in range(max(1, restarts)):
		# 乱数をずらす（seed未指定なら毎回異なる）
		sub_seed = base_rnd.getrandbits(64)
		order = diversified_order(participants, sub_seed)

		# 1) 先読みで割付
		rnd = random.Random(sub_seed)
		pairs = _assign_pairs_with_bye_and_lookahead(order, M, bye_pos, lookahead, rnd)

		c1, c2, c3 = _conflicts(pairs)
		c4 = _half_distribution_penalty(pairs)
		c5 = _future_round_penalty(pairs)
		
		if best_pairs == None or (c1, c2, c3, c4, c5) < (best_c1, best_c2, best_c3, best_c4, best_c5):
			best_pairs, best_c1, best_c2, best_c3, best_c4, best_c5 = pairs, c1, c2, c3, c4, c5
			if best_c1 == 0 and best_c2 == 0 and best_c3 == 0 and best_c4 == 0 and best_c5 == 0:   # これ以上はない
				break

	print(f"c1={best_c1}, c2={best_c2}, c3={best_c3}, c4={best_c4}, c5={best_c5}")
	return best_pairs if best_pairs is not None else []

# --------------------
# ブラケット構築（決勝まで）
# --------------------
def build_full_bracket(participants: List[Participant], seed: Optional[int] = None) -> List[Round]:
	first_pairs = make_first_round_pairs_quarter_even(participants, seed=seed, restarts=1000)
	rounds: List[Round] = []

	r0: Round = []
	for (a, b) in first_pairs:
		m = Match(left=a, right=b, winner=None)
		# BYE自動処理（1回戦のみ）
		if a.name == "BYE" and b.name != "BYE":
			m.winner = b
		elif b.name == "BYE" and a.name != "BYE":
			m.winner = a
		elif a.name == "BYE" and b.name == "BYE":
			# ここは起きない想定（禁止済み）だが保険で右勝者
			m.winner = b
		r0.append(m)
	rounds.append(r0)

	current_round = r0
	while len(current_round) > 1:
		next_round: Round = []
		for i in range(0, len(current_round), 2):
			L, R = current_round[i], current_round[i + 1]
			left_slot  = L.winner if L.winner else Ref(i)
			right_slot = R.winner if R.winner else Ref(i + 1)
			next_round.append(Match(left=left_slot, right=right_slot, winner=None))
		rounds.append(next_round)
		current_round = next_round
	return rounds

# ----------------------------
# コンソール出力ユーティリティ
# ----------------------------
def print_bracket(rounds: List[Round]) -> None:
	def round_name(idx: int, total_rounds: int) -> str:
		rnum = idx + 1
		if rnum == total_rounds: return "Final"
		if rnum == total_rounds - 1: return "Semifinal"
		if rnum == total_rounds - 2: return "Quarterfinal"
		return f"Round {rnum}"

	def slot_str(slot: Slot) -> str:
		if isinstance(slot, Participant):
			return f"{slot.name} ({slot.groupname or '-'})"
		return f"Winner of Match {slot.match_index + 1}"

	total = len(rounds)
	for r_idx, rnd in enumerate(rounds):
		#print(f"\n=== {round_name(r_idx, total)} ===")
		for m_idx, m in enumerate(rnd, start=1):
			left = slot_str(m.left)
			right = slot_str(m.right)
			#w = f" -> WINNER: {m.winner.name}" if m.winner else ""
			w = ""
			print(f"Match {m_idx}: {left}  vs  {right}{w}")
		break

# ------------------------------
# Excelグラフィカル出力ユーティリティ
# ------------------------------
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Side

def save_bracket_xlsx(
	rounds: List[Round],
	path="bracket.xlsx",
	init_workbook = False,
	hide_groupname = False, # 団体名称の列を隠すかどうか。個人戦の場合は隠さず表示するためFalse、団体戦トーナメントの場合は選手名称に団体名が含まれるのでTrueにして隠す。
	summary_name = "xxxxxxxxの部", # シート名
	match_name = "xxxxxxxxの部", # 部門名称
	match_date = "2099.12.31",
	match_place1 = "第一試合場",
	match_place2 = "第二試合場",
):
	# ─────────────────────────────
	#  Excel 初期化
	# ─────────────────────────────
	if init_workbook:
		# 既存のファイルを削除し、新しいエクセルファイルを作成する
		try:
			os.remove(path)
			print(f"[{path}] を削除しました。")
		except FileNotFoundError:
			pass
		wb = Workbook()

		# 新しいシート{summary_name}を作成し、それ以外のデフォルトのシートを削除する
		ws = wb.create_sheet(summary_name)
		for sheet_name in wb.sheetnames:
			if sheet_name != summary_name:
				wb.remove(wb[sheet_name])
		wb.save(path)
		print(f"[{path}] を作成しました。")

	else:
		# 既存のエクセルファイルを開く
		wb = load_workbook(path)

		# 既存のシート{summary_name}を削除し、新しいシート{summary_name}を作成する
		if summary_name in  wb.sheetnames:
			wb.remove(wb[summary_name])
		ws = wb.create_sheet(summary_name)
		wb.save(path)

	ws.title = summary_name
	ws.column_dimensions["A"].width = ws.column_dimensions["A"].width / 4 # 一番左の列幅
	ws.column_dimensions["B"].width = ws.column_dimensions["B"].width / 3 # 選手番号の列幅
	ws.column_dimensions["C"].width = ws.column_dimensions["C"].width * 2 # 選手名称の列幅
	ws.column_dimensions["D"].width = ws.column_dimensions["D"].width * 1.25 # 団体名称の列幅
	if hide_groupname == True:
		ws.column_dimensions["D"].hidden = True

	thin = Side(style="thin") # トーナメント線のスタイル

	# ─────────────────────────────
	#  helpers
	# ─────────────────────────────
	def is_real(x: Slot) -> bool:
		return isinstance(x, Participant) and x.name != "BYE"

	def assign_match_numbers(rounds: List[Round]) -> dict[tuple[int, int], int]:
		"""BYEを除く全ラウンドの実施試合に通し番号を振る。"""
		n = 1
		mapping = {}
		for r, rnd in enumerate(rounds):
			for i, m in enumerate(rnd):
				# 左右のどちらも BYE でない（＝実際に行われる試合）
				left_is_bye  = isinstance(m.left, Participant)  and m.left.name  == "BYE"
				right_is_bye = isinstance(m.right, Participant) and m.right.name == "BYE"
				if not (left_is_bye or right_is_bye):
					mapping[(r, i)] = n
					n += 1
		return mapping

	match_no = assign_match_numbers(rounds)

	R = len(rounds)

	# ─────────────────────────────
	#  部門情報
	# ─────────────────────────────
	row = 1
	ws.cell(row, 1).value = match_name
	row += 1
	ws.cell(row, 2).value = match_date
	row += 1
	ws.cell(row, 2).value = match_place1
	row += 1
	ws.cell(row, 2).value = match_place2

	# ─────────────────────────────
	#  "1回戦", ... "準決勝", "決勝" のラベル出力
	# ─────────────────────────────
	row += 1
	for r in range(0, R):
		if r == R - 1:
			ws.cell(row, r + 5).value = "決勝"
		elif r == R - 2:
			ws.cell(row, r + 5).value = "準決勝"
		else:
			ws.cell(row, r + 5).value = f"{r + 1}回戦"
		ws.cell(row, r + 5).alignment = Alignment(horizontal="right")

	# ─────────────────────────────
	#  選手情報
	# ─────────────────────────────
	row_player_1st = row + 1 # 最初の選手の行を保存しておく

	# 選手番号のマップを作成
	player_numbering: dict[Participant, int] = {}

	# 選手の行番号マップを作製
	player_row: dict[Participant, int] = {}
	
	player_counter = 0
	r0 = rounds[0]
	for match in r0:
		for player in (match.left, match.right):
			if is_real(player):
				player_counter += 1
				player_numbering[player] = player_counter
				player_row[player] = row_player_1st + (player_counter - 1) * 2

	def output_player(player):
		player_no = player_numbering.get(player)
		row = row_player_1st + (player_no - 1) * 2

		# セル結合
		ws.merge_cells(start_row=row, start_column=2, end_row=row+1, end_column=2)
		ws.merge_cells(start_row=row, start_column=3, end_row=row+1, end_column=3)
		ws.merge_cells(start_row=row, start_column=4, end_row=row+1, end_column=4)

		# 文字位置調整
		ws.cell(row, 2).alignment = Alignment(horizontal="right", vertical="center")
		ws.cell(row, 3).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
		ws.cell(row, 4).alignment = Alignment(horizontal="left", vertical="center")

		# 縦幅を拡張し、選手と選手の間にスペースを設ける（縦幅のデフォルトは約15らしい）
		ws.row_dimensions[row].height = 15 * 1.25
		ws.row_dimensions[row+1].height = 15 * 1.25

		# 選手情報出力
		ws.cell(row, 2).value = player_no
		kana = f"\n({player.kana})" if player.kana != None and player.kana != "" else ""
		ws.cell(row, 3).value = player.name + kana
		ws.cell(row, 4).value = f"[{player.groupname}]"

	for match in r0:
		if is_real(match.left):
			output_player(match.left)

		if is_real(match.right):
			output_player(match.right)

	# ─────────────────────────────
	#  トーナメント線、試合番号
	# ─────────────────────────────
	match_2_row = {} # KEY:match、VALUE:match勝者の水平線の直上行番号

	for rnd_idx, rnd in enumerate(rounds):
		rnd_no = rnd_idx + 1
		rnd_col = rnd_no + 4
		
		# --- 1回戦 ---
		if rnd_no == 1:
			for match_idx, match in enumerate(rnd):
				if is_real(match.left):
					# 選手は2行で構成されるが、その1行目の底辺に水平線を引く
					row = player_row.get(match.left)
					ws.cell(row, rnd_col).border = Border(bottom=thin)
				
				if is_real(match.right):
					# 選手は2行で構成されるが、その2行目の上辺に水平線を引く
					row = player_row.get(match.right)
					ws.cell(row + 1, rnd_col).border = Border(top=thin)
				
				if is_real(match.left) and is_real(match.right):
					# 二人の選手の水平線に囲まれたセルを結合し、右辺に垂直線を引く
					row_left = player_row.get(match.left)
					row_right = player_row.get(match.right)
					ws.merge_cells(start_row=row_left+1, start_column=rnd_col, end_row=row_right, end_column=rnd_col) # 結合
					for row in range(row_left+1, row_right + 1):
						ws.cell(row, rnd_col).border = Border(right=thin) # 垂直線
					
					# 試合番号
					ws.cell(row_left + 1, rnd_col).value = match_no.get((rnd_idx, match_idx))
					ws.cell(row_left + 1, rnd_col).alignment = Alignment(horizontal="right", vertical="center")
				
				# match勝者の水平線の直上行番号を記録する
				if is_real(match.left) and is_real(match.right):
					match_2_row[(rnd_idx, match_idx)] = player_row.get(match.left) + 1
				elif is_real(match.left):
					match_2_row[(rnd_idx, match_idx)] = player_row.get(match.left)
				elif is_real(match.right):
					match_2_row[(rnd_idx, match_idx)] = player_row.get(match.right)

		# --- 2回戦～決勝 ---
		else:
			for match_idx, match in enumerate(rnd):
				# 選手1の行の底辺に水平線を引く
				row_left = match_2_row[(rnd_idx - 1, match_idx * 2)]
				ws.cell(row_left, rnd_col).border = Border(bottom=thin)

				# 選手2の行+1の上辺に水平線を引く
				row_right = match_2_row[(rnd_idx - 1, match_idx * 2 + 1)]
				ws.cell(row_right + 1, rnd_col).border = Border(top=thin)

				# 二人の選手の水平線に囲まれたセルを結合し、右辺に垂直線を引く
				ws.merge_cells(start_row=row_left+1, start_column=rnd_col, end_row=row_right, end_column=rnd_col) # 結合
				for row in range(row_left+1, row_right + 1):
					ws.cell(row, rnd_col).border = Border(right=thin) # 垂直線

				# 試合番号
				ws.cell(row_left + 1, rnd_col).value = match_no.get((rnd_idx, match_idx))
				ws.cell(row_left + 1, rnd_col).alignment = Alignment(horizontal="right", vertical="center")

				# match勝者の水平線の直上行番号を記録する
				match_2_row[(rnd_idx, match_idx)] = (row_left + row_right) // 2

	# 決勝戦の垂直線の中心から右に水平線を引く
	rnd_col = 4 + len(rounds) + 1
	ws.cell(match_2_row[len(rounds) - 1, 0], rnd_col).border = Border(bottom=thin)

	# ファイルを保存する
	wb.save(path)


# --------------------
# サンプル実行
# --------------------
if __name__ == "__main__":
	sample = [
		#Participant("伊藤　創", "いとう　はじめ", "中央会"),
		#Participant("荒井　裕幸", "あらい　ひろゆき", "中央会"),
		#Participant("柳田　哲朗", "やなぎだ　てつお", "中央会"),
		#Participant("齋藤　直樹", "さいとう　なおき", "中央会"),
		#Participant("石黒　健司", "いしぐろ　けんじ", "中央会"),
		#Participant("塩幡　勝典", "しおはた　かつのり", "中央会"),
		#Participant("杉野　仁司", "すぎの　ひとし", "中央会"),
		#Participant("福田　知広", "ふくだ　かずひろ", "中央会"),
		#Participant("北村 和也", "きたむら　かずや", "大町"),
		#Participant("富宇加　健", "とみうか　けん", "聖武会"),
		#Participant("中嶋　昭弘", "なかじま　あきひろ", "深大寺"),
		#Participant("磯野　渉", "いその　わたる", "深大寺"),
		#Participant("松本　義弘", "まつもと　よしひろ", "深大寺"),
		#Participant("國分　崇生", "こくぶん　たかお", "電通大"),
		
		#Participant("齋藤　美亜希", "さいとう　みつき", "中央会"),
		#Participant("上渡　勇人", "かみわたり　はやと", "中央会"),
		#Participant("塩幡　結依子", "しおはた　ゆいこ", "中央会"),
		#Participant("吉岡　侑哉", "よしおか　ゆうや", "大町"),
		#Participant("井上　公稀", "いのうえ　こうき", "聖武会"),
		#Participant("冨宇加　育実", "とみうか　いくみ", "聖武会"),
		#Participant("堂崎　菖瑚", "どうさき　しょうご", "聖武会"),
		#Participant("大塚　晴天", "おおつか　はるたか", "文荘館"),
		#Participant("土屋　美穏", "つちや　みお", "第七機動隊"),
		#Participant("伊藤　朱嶺", "いとう　あかね", "第七機動隊"),
		
		Participant("柳沢　信高", "やなぎさわ　のぶたか", "中央会"),
		Participant("大濵　賢吾", "おおはま　けんご", "中央会"),
		Participant("大濵　賢汰", "おおはま　けんた", "中央会"),
		Participant("筒井　貫太", "つつい　かんた", "中央会"),
		#Participant("五十嵐  周作", "いがらし　しゅうさく", "中央会"),
		Participant("眞野　仁", "まの　ひとし", "染地"),
		Participant("東條　剛", "とうじょう　つよし", "染地"),
		Participant("松浦　舜穏", "まつうら　しおん", "染地"),
		Participant("三石　利明", "みついし　としあき", "染地"),
		Participant("渡邉　正人", "わたなべ　まさと", "染地"),
		Participant("佐藤　和", "さとう　やまと", "染地"),
		Participant("五十畑　伊織", "いそはた　いおり", "大町"),
		Participant("大武　凪希", "おおたけ　なぎ", "聖武会"),
		Participant("大武　正治", "おおたけ　まさはる", "聖武会"),
		Participant("内田　佑樹", "うちだ　ゆうき", "聖武会"),
		Participant("寺田　直人", "てらだ　なおと", "聖武会"),
		Participant("堤　文彦", "つつみ　ふみひこ", "聖武会"),
		Participant("吉川　陽色", "よしかわ　ひいろ", "聖武会"),
		Participant("大岡　克也", "おおおか　かつや", "深大寺"),
		Participant("渡邉　英明", "わたなべ　ひであき", "深大寺"),
		Participant("小林　泰宏", "こばやし　やすひろ", "深大寺"),
		Participant("廣井　騰哉", "ひろい　とうや", "深大寺"),
		Participant("神山　遼太郎", "かみやま　りょうたろう", "電通大"),
		Participant("木村　琉人", "きむら　りゅうと", "電通大"),
		Participant("増澤　日路", "ますざわ　ひろ", "電通大"),
		Participant("鈴木　康太", "すずき　こうた", "電通大"),
		Participant("清水　陽平", "しみず　ようへい", "電通大"),
		Participant("井上　勇気", "いのうえ　ゆうき", "電通大"),
		Participant("名久井　龍太郎", "なくい　りゅうたろう", "電通大"),
		Participant("佐藤　海渡", "さとう　かいと", "電通大"),
		Participant("永野　照幸", "ながの　てるゆき", "狛江高"),
		Participant("仲　芳弘", "なか　よしひろ", "個人参加"),
		Participant("冨田　大介", "とみた　だいすけ", "個人参加"),

		#Participant("2柳沢　信高", "やなぎさわ　のぶたか", "中央会"),
		#Participant("2大濵　賢吾", "おおはま　けんご", "中央会"),
		#Participant("2大濵　賢汰", "おおはま　けんた", "中央会"),
		#Participant("2筒井　貫太", "つつい　かんた", "中央会"),
		#Participant("2五十嵐  周作", "いがらし　しゅうさく", "中央会"),
		#Participant("2眞野　仁", "まの　ひとし", "染地"),
		#Participant("2東條　剛", "とうじょう　つよし", "染地"),
		#Participant("2松浦　舜穏", "まつうら　しおん", "染地"),
		#Participant("2三石　利明", "みついし　としあき", "染地"),
		#Participant("2渡邉　正人", "わたなべ　まさと", "染地"),
		#Participant("2佐藤　和", "さとう　やまと", "染地"),
		#Participant("2五十畑　伊織", "いそはた　いおり", "大町"),
		#Participant("2大武　凪希", "おおたけ　なぎ", "聖武会"),
		#Participant("2大武　正治", "おおたけ　まさはる", "聖武会"),
		#Participant("2内田　佑樹", "うちだ　ゆうき", "聖武会"),
		#Participant("2寺田　直人", "てらだ　なおと", "聖武会"),
		#Participant("2堤　文彦", "つつみ　ふみひこ", "聖武会"),
		#Participant("2吉川　陽色", "よしかわ　ひいろ", "聖武会"),
		#Participant("2大岡　克也", "おおおか　かつや", "深大寺"),
		#Participant("2渡邉　英明", "わたなべ　ひであき", "深大寺"),
		#Participant("2小林　泰宏", "こばやし　やすひろ", "深大寺"),
		#Participant("2廣井　騰哉", "ひろい　とうや", "深大寺"),
		#Participant("2神山　遼太郎", "かみやま　りょうたろう", "電通大"),
		#Participant("2木村　琉人", "きむら　りゅうと", "電通大"),
		#Participant("2増澤　日路", "ますざわ　ひろ", "電通大"),
		#Participant("2鈴木　康太", "すずき　こうた", "電通大"),
		#Participant("2清水　陽平", "しみず　ようへい", "電通大"),
		#Participant("2井上　勇気", "いのうえ　ゆうき", "電通大"),
		#Participant("2名久井　龍太郎", "なくい　りゅうたろう", "電通大"),
		#Participant("2佐藤　海渡", "さとう　かいと", "電通大"),
		#Participant("2永野　照幸", "ながの　てるゆき", "狛江高"),
		#Participant("2仲　芳弘", "なか　よしひろ", "個人参加"),
		#Participant("2冨田　大介", "とみた　だいすけ", "個人参加"),
		#
		#Participant("3柳沢　信高", "やなぎさわ　のぶたか", "中央会"),
		#Participant("3大濵　賢吾", "おおはま　けんご", "中央会"),
		#Participant("3大濵　賢汰", "おおはま　けんた", "中央会"),
		#Participant("3筒井　貫太", "つつい　かんた", "中央会"),
		#Participant("3五十嵐  周作", "いがらし　しゅうさく", "中央会"),
		#Participant("3眞野　仁", "まの　ひとし", "染地"),
		#Participant("3東條　剛", "とうじょう　つよし", "染地"),
		#Participant("3松浦　舜穏", "まつうら　しおん", "染地"),
		#Participant("3三石　利明", "みついし　としあき", "染地"),
		#Participant("3渡邉　正人", "わたなべ　まさと", "染地"),
		#Participant("3佐藤　和", "さとう　やまと", "染地"),
		#Participant("3五十畑　伊織", "いそはた　いおり", "大町"),
		#Participant("3大武　凪希", "おおたけ　なぎ", "聖武会"),
		#Participant("3大武　正治", "おおたけ　まさはる", "聖武会"),
		#Participant("3内田　佑樹", "うちだ　ゆうき", "聖武会"),
		#Participant("3寺田　直人", "てらだ　なおと", "聖武会"),
		#Participant("3堤　文彦", "つつみ　ふみひこ", "聖武会"),
		#Participant("3吉川　陽色", "よしかわ　ひいろ", "聖武会"),
		#Participant("3大岡　克也", "おおおか　かつや", "深大寺"),
		#Participant("3渡邉　英明", "わたなべ　ひであき", "深大寺"),
		#Participant("3小林　泰宏", "こばやし　やすひろ", "深大寺"),
		#Participant("3廣井　騰哉", "ひろい　とうや", "深大寺"),
		#Participant("3神山　遼太郎", "かみやま　りょうたろう", "電通大"),
		#Participant("3木村　琉人", "きむら　りゅうと", "電通大"),
		#Participant("3増澤　日路", "ますざわ　ひろ", "電通大"),
		#Participant("3鈴木　康太", "すずき　こうた", "電通大"),
		#Participant("3清水　陽平", "しみず　ようへい", "電通大"),
		#Participant("3井上　勇気", "いのうえ　ゆうき", "電通大"),
		#Participant("3名久井　龍太郎", "なくい　りゅうたろう", "電通大"),
		#Participant("3佐藤　海渡", "さとう　かいと", "電通大"),
		#Participant("3永野　照幸", "ながの　てるゆき", "狛江高"),
		#Participant("3仲　芳弘", "なか　よしひろ", "個人参加"),
		#Participant("3冨田　大介", "とみた　だいすけ", "個人参加"),
	]
	
	#rounds = build_full_bracket(sample, seed=1)
	rounds = build_full_bracket(sample, seed=20251026 * 10 + 1)
	
	print_bracket(rounds)
	#save_bracket_svg(rounds, "result.svg")
	save_bracket_xlsx(rounds, "result.xlsx", match_name="一般六･七段の部", match_date="2099.12.31", match_place1="第一試合場(選手番号1～7)", match_place2="第二試合場(選手番号8～14)", init_workbook=True)
